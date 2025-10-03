#!/usr/bin/env /Users/leechanhee/ConstructionManagement-Installer/venv/bin/python
# -*- coding: utf-8 -*-
"""
청구서 템플릿 생성기
기존 템플릿을 플레이스홀더 방식으로 수정합니다.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def create_invoice_template():
    """플레이스홀더 기반 청구서 템플릿을 생성합니다."""
    
    # 기존 템플릿 로드
    template_path = "/Users/leechanhee/ConstructionManagement-Installer/docs/청구서 상세 폼_backup.xlsx"
    output_path = "/Users/leechanhee/ConstructionManagement-Installer/docs/청구서 상세 폼.xlsx"
    
    wb = openpyxl.load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]
    
    print("기존 템플릿 로드 완료")
    
    # 1) 제목은 그대로 유지
    # A1:I1 "청구서 상세 -" 그대로
    
    # 2) 헤더 정보를 플레이스홀더로 수정
    # A3:I3 건축주
    ws.cell(row=3, column=1).value = "건 축 주 : {client}"
    
    # A4:I4 프로젝트  
    ws.cell(row=4, column=1).value = "프로젝트 : {project}"
    
    # A5:I5 작업장 주소
    ws.cell(row=5, column=1).value = "작업장 주소 : {site_addr}"
    
    # 발행일 설정 (오른쪽에 배치)
    # 기존 병합 해제 후 수정
    try:
        # AK3:AP3 영역 병합 해제
        ws.unmerge_cells('AK3:AP3')
    except:
        pass
    
    # 발행일 설정
    ws.cell(row=3, column=37).value = "발행일 : {issued_at}"
    # 다시 병합
    ws.merge_cells('AK3:AP3')
    
    print("헤더 플레이스홀더 설정 완료")
    
    # 3) 항목 반복 영역 설정
    # 7행 다음에 반복 마커 추가
    
    # 새로운 행들을 삽입
    ws.insert_rows(8, 3)  # 8행에 3개 행 삽입
    
    # 시작 마커 (8행)
    ws.cell(row=8, column=1).value = "{#items}"
    
    # 템플릿 행 (9행) - 실제 데이터 구조
    template_row = 9
    
    # A9:J9 "내용" - {item.title}\n{item.desc}
    ws.merge_cells(f'A{template_row}:J{template_row}')
    content_cell = ws.cell(row=template_row, column=1)
    content_cell.value = "{item.title}\n{item.desc}"
    content_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    content_cell.border = Border(
        top=Side(style='thin'), bottom=Side(style='thin'),
        left=Side(style='thin'), right=Side(style='thin')
    )
    
    # K9:P9 "규격" - {item.spec}
    ws.merge_cells(f'K{template_row}:P{template_row}')
    spec_cell = ws.cell(row=template_row, column=11)
    spec_cell.value = "{item.spec}"
    spec_cell.alignment = Alignment(horizontal='center', vertical='center')
    spec_cell.border = Border(
        top=Side(style='thin'), bottom=Side(style='thin'),
        left=Side(style='thin'), right=Side(style='thin')
    )
    
    # Q9:U9 "수량" - {item.qty}
    ws.merge_cells(f'Q{template_row}:U{template_row}')
    qty_cell = ws.cell(row=template_row, column=17)
    qty_cell.value = "{item.qty}"
    qty_cell.alignment = Alignment(horizontal='center', vertical='center')
    qty_cell.border = Border(
        top=Side(style='thin'), bottom=Side(style='thin'),
        left=Side(style='thin'), right=Side(style='thin')
    )
    
    # V9:Z9 "단위" - {item.unit}
    ws.merge_cells(f'V{template_row}:Z{template_row}')
    unit_cell = ws.cell(row=template_row, column=22)
    unit_cell.value = "{item.unit}"
    unit_cell.alignment = Alignment(horizontal='center', vertical='center')
    unit_cell.border = Border(
        top=Side(style='thin'), bottom=Side(style='thin'),
        left=Side(style='thin'), right=Side(style='thin')
    )
    
    # AA9:AG9 "단가" - {item.unit_price}
    ws.merge_cells(f'AA{template_row}:AG{template_row}')
    price_cell = ws.cell(row=template_row, column=27)
    price_cell.value = "{item.unit_price}"
    price_cell.alignment = Alignment(horizontal='right', vertical='center')
    price_cell.number_format = "#,##0"
    price_cell.border = Border(
        top=Side(style='thin'), bottom=Side(style='thin'),
        left=Side(style='thin'), right=Side(style='thin')
    )
    
    # AH9:AO9 "합계" - 수식으로 계산
    ws.merge_cells(f'AH{template_row}:AO{template_row}')
    total_cell = ws.cell(row=template_row, column=34)
    total_cell.value = "=Q9*AA9"  # 수량 * 단가
    total_cell.alignment = Alignment(horizontal='right', vertical='center')
    total_cell.number_format = "#,##0"
    total_cell.border = Border(
        top=Side(style='thin'), bottom=Side(style='thin'),
        left=Side(style='thin'), right=Side(style='thin')
    )
    
    # AP9:AZ9 "비고" - {item.note}
    ws.merge_cells(f'AP{template_row}:AZ{template_row}')
    note_cell = ws.cell(row=template_row, column=42)
    note_cell.value = "{item.note}"
    note_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    note_cell.border = Border(
        top=Side(style='thin'), bottom=Side(style='thin'),
        left=Side(style='thin'), right=Side(style='thin')
    )
    
    # 템플릿 행 높이 설정
    ws.row_dimensions[template_row].height = 30
    
    # 종료 마커 (10행)
    ws.cell(row=10, column=1).value = "{/items}"
    
    print("템플릿 행 및 마커 설정 완료")
    
    # 4) 총합계 행 수정 (원래 9행이었지만 이제 11행)
    total_sum_row = 11
    
    # 기존 총합계 행의 내용을 수정
    # A11:AO11 "총 합계 :"
    ws.merge_cells(f'A{total_sum_row}:AO{total_sum_row}')
    total_label_cell = ws.cell(row=total_sum_row, column=1)
    total_label_cell.value = "총 합계 :"
    total_label_cell.font = Font(bold=True, size=12)
    total_label_cell.alignment = Alignment(horizontal='left', vertical='center')
    total_label_cell.border = Border(
        top=Side(style='thick'), bottom=Side(style='thick'),
        left=Side(style='thin'), right=Side(style='thin')
    )
    
    # AP11:AZ11 총합계 금액 - {TOTAL_SUM} 플레이스홀더
    ws.merge_cells(f'AP{total_sum_row}:AZ{total_sum_row}')
    total_amount_cell = ws.cell(row=total_sum_row, column=42)
    total_amount_cell.value = "{TOTAL_SUM}"
    total_amount_cell.font = Font(bold=True, size=12)
    total_amount_cell.alignment = Alignment(horizontal='right', vertical='center')
    total_amount_cell.fill = PatternFill(start_color='FFFFEB3B', end_color='FFFFEB3B', fill_type='solid')
    total_amount_cell.border = Border(
        top=Side(style='thick'), bottom=Side(style='thick'),
        left=Side(style='thin'), right=Side(style='thin')
    )
    
    print("총합계 플레이스홀더 설정 완료")
    
    # 5) 파일 저장
    wb.save(output_path)
    print(f"플레이스홀더 템플릿 생성 완료: {output_path}")
    
    return output_path

if __name__ == "__main__":
    create_invoice_template()