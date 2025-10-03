#!/usr/bin/env /Users/leechanhee/ConstructionManagement-Installer/venv/bin/python
# -*- coding: utf-8 -*-
"""
청구서 템플릿 수정
사용자가 수정한 템플릿을 분석하고 누락된 플레이스홀더를 추가합니다.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

def fix_template():
    """사용자가 수정한 템플릿을 분석하고 수정합니다."""
    
    template_path = "/Users/leechanhee/ConstructionManagement-Installer/docs/청구서 상세 폼.xlsx"
    
    wb = openpyxl.load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]
    
    print("템플릿 수정 시작...")
    
    # 현재 9행의 병합을 해제
    try:
        ws.unmerge_cells('A9:AO9')
        print("9행 병합 해제 완료")
    except:
        print("9행 병합 해제 실패 또는 이미 해제됨")
    
    # 각 컬럼별로 플레이스홀더 설정
    template_row = 9
    
    # A9:J9 "내용" - {item.title}\n{item.desc}
    ws.merge_cells(f'A{template_row}:J{template_row}')
    content_cell = ws.cell(row=template_row, column=1)
    content_cell.value = "{item.title}\n{item.desc}"
    content_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # K9:P9 "규격" - {item.spec}
    ws.merge_cells(f'K{template_row}:P{template_row}')
    spec_cell = ws.cell(row=template_row, column=11)
    spec_cell.value = "{item.spec}"
    spec_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Q9:U9 "수량" - {item.qty}
    ws.merge_cells(f'Q{template_row}:U{template_row}')
    qty_cell = ws.cell(row=template_row, column=17)
    qty_cell.value = "{item.qty}"
    qty_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # V9:Z9 "단위" - {item.unit}
    ws.merge_cells(f'V{template_row}:Z{template_row}')
    unit_cell = ws.cell(row=template_row, column=22)
    unit_cell.value = "{item.unit}"
    unit_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # AA9:AG9 "단가" - {item.unit_price}
    ws.merge_cells(f'AA{template_row}:AG{template_row}')
    price_cell = ws.cell(row=template_row, column=27)
    price_cell.value = "{item.unit_price}"
    price_cell.alignment = Alignment(horizontal='right', vertical='center')
    price_cell.number_format = "#,##0"
    
    # AH9:AO9 "합계" - 수식으로 계산
    ws.merge_cells(f'AH{template_row}:AO{template_row}')
    total_cell = ws.cell(row=template_row, column=34)
    total_cell.value = "=Q9*AA9"  # 수량 * 단가
    total_cell.alignment = Alignment(horizontal='right', vertical='center')
    total_cell.number_format = "#,##0"
    
    # AP9:AZ9 "비고" - {item.note} (이미 있음)
    ws.merge_cells(f'AP{template_row}:AZ{template_row}')
    note_cell = ws.cell(row=template_row, column=42)
    note_cell.value = "{item.note}"
    note_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 모든 셀에 테두리 추가
    thin_border = Border(
        top=Side(style='thin'), 
        bottom=Side(style='thin'),
        left=Side(style='thin'), 
        right=Side(style='thin')
    )
    
    for c in range(1, 53):  # 모든 열
        cell = ws.cell(row=template_row, column=c)
        cell.border = thin_border
    
    # 템플릿 행 높이 설정
    ws.row_dimensions[template_row].height = 30
    
    print("템플릿 플레이스홀더 수정 완료")
    
    # 파일 저장
    wb.save(template_path)
    print(f"수정된 템플릿 저장 완료: {template_path}")
    
    return template_path

if __name__ == "__main__":
    fix_template()