#!/usr/bin/env /Users/leechanhee/ConstructionManagement-Installer/venv/bin/python
# -*- coding: utf-8 -*-
"""
청구서 템플릿 렌더러
openpyxl을 사용하여 플레이스홀더 기반 템플릿을 렌더링합니다.
"""

import sys
import json
import argparse
from copy import copy
from pathlib import Path
import re
import openpyxl
from openpyxl.styles import Alignment, Font

# ---------- 유틸 함수 ----------
def get_value_by_path(data, path_str):
    """점표기 경로(header.client 등) 해석"""
    cur = data
    for p in path_str.split("."):
        if isinstance(cur, dict) and p in cur:
            cur = cur[p]
        else:
            return None
    return cur

def replace_placeholders_in_text(text, payload):
    """문자열 안의 {…} 패턴들을 payload 값으로 부분 치환"""
    def repl(m):
        key = m.group(1).strip()
        val = get_value_by_path(payload, key)
        return "" if val is None else str(val)
    return re.sub(r"\{([^{}]+)\}", repl, text)

def clone_cell_style(src, dst):
    """셀 스타일을 복제합니다."""
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.number_format = copy(src.number_format)
    dst.protection = copy(src.protection)
    dst.alignment = copy(src.alignment)
    dst.border = copy(src.border)

def horizontal_merges_for_row(ws, row):
    """해당 행의 수평 병합 범위를 반환합니다."""
    merges = []
    for rng in ws.merged_cells.ranges:
        if rng.min_row == row and rng.max_row == row:
            merges.append((rng.min_col, rng.max_col))
    merges.sort()
    return merges

def apply_horizontal_merges(ws, merges, target_row):
    """수평 병합 범위를 target_row에 적용합니다."""
    for c1, c2 in merges:
        if c1 < c2:
            ws.merge_cells(start_row=target_row, start_column=c1, end_row=target_row, end_column=c2)

# ---------- 메인 렌더링 함수 ----------
def render_invoice(template_path, output_path, payload):
    """템플릿을 렌더링하여 청구서를 생성합니다."""
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb[wb.sheetnames[0]]

        # 디버그 메시지를 stderr로 출력
        print(f"템플릿 로드 완료: {template_path}", file=sys.stderr)
        print(f"워크시트: {ws.title}, 최대 행: {ws.max_row}, 최대 열: {ws.max_column}", file=sys.stderr)

        # 1) 전역 플레이스홀더 치환(모든 셀 순회) - 반복 마커는 제외
        print("1단계: 전역 플레이스홀더 치환 중...", file=sys.stderr)
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, str) and "{" in cell.value and "}" in cell.value:
                    # 반복 마커와 item 플레이스홀더는 치환하지 않음
                    if "{#items}" in cell.value or "{/items}" in cell.value or "{item." in cell.value:
                        continue
                    old_value = cell.value
                    new_value = replace_placeholders_in_text(cell.value, payload)
                    if old_value != new_value:
                        print(f"  치환: {old_value} → {new_value}", file=sys.stderr)
                        cell.value = new_value

        # 2) 항목 반복 블록 찾기
        print("2단계: 항목 반복 블록 찾기...", file=sys.stderr)
        start_row = end_row = None
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell_value = str(ws.cell(row=r, column=c).value or "")
                if "{#items}" in cell_value:
                    start_row = r
                    print(f"  시작 행 발견: {r}행 {c}열", file=sys.stderr)
                elif "{/items}" in cell_value:
                    end_row = r
                    print(f"  종료 행 발견: {r}행 {c}열", file=sys.stderr)
                    break
            if end_row:
                break

        if start_row is None or end_row is None or end_row <= start_row + 1:
            print("  반복 블록이 없습니다. 기본 렌더링 완료.", file=sys.stderr)
            wb.save(output_path)
            return {"success": True, "output_path": str(output_path)}

        template_row = start_row + 1  # 이 한 줄이 항목 템플릿
        items = payload.get("items", [])
        print(f"  템플릿 행: {template_row}, 항목 수: {len(items)}", file=sys.stderr)

        # 템플릿 행 스타일/병합/행높이 보관
        max_cols = ws.max_column
        tmpl_cells = [ws.cell(row=template_row, column=c) for c in range(1, max_cols + 1)]
        tmpl_height = ws.row_dimensions[template_row].height
        tmpl_merges = horizontal_merges_for_row(ws, template_row)
        print(f"  템플릿 행 높이: {tmpl_height}, 병합 구조: {tmpl_merges}", file=sys.stderr)

        # 3) 템플릿 아래에 items 길이만큼 공간 확보
        if items and len(items) > 1:
            print(f"3단계: {len(items) - 1}개 행 삽입...", file=sys.stderr)
            ws.insert_rows(template_row + 1, amount=len(items) - 1)

        # 4) 각 항목 렌더링
        print("4단계: 항목 데이터 렌더링...", file=sys.stderr)
        first_data_row = template_row
        for i, item in enumerate(items):
            r = first_data_row + i
            print(f"  항목 {i + 1} 렌더링 (행 {r}): {item.get('title', 'N/A')}", file=sys.stderr)
            
            # 동일 병합 구조 적용
            apply_horizontal_merges(ws, tmpl_merges, r)

            # 셀 값 작성(템플릿 셀을 문자열 치환한 결과를 사용)
            for c in range(1, max_cols + 1):
                src = tmpl_cells[c - 1]
                dst = ws.cell(row=r, column=c, value=src.value)
                clone_cell_style(src, dst)
                
                if isinstance(dst.value, str) and "{item." in dst.value:
                    # {item.키} 치환
                    old_value = dst.value
                    dst.value = re.sub(
                        r"\{item\.([^{}]+)\}",
                        lambda m: str(item.get(m.group(1).strip(), "")),
                        dst.value
                    )
                    if old_value != dst.value:
                        print(f"    항목 {i + 1} 치환 (열 {c}): {old_value} → {dst.value}", file=sys.stderr)

            # 수식/숫자형식 적용 (필요시)
            # 수량, 단가, 합계 열을 찾아서 적절히 설정
            for c in range(1, max_cols + 1):
                cell = ws.cell(row=r, column=c)
                # 숫자 셀에 천단위 콤마 형식 적용
                if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.isdigit()):
                    cell.number_format = "#,##0"

            # 행 높이 설정
            if tmpl_height:
                ws.row_dimensions[r].height = tmpl_height

        # 5) 템플릿/마커 행 삭제({#items}, {/items}, 템플릿 행)
        #    주의: 아래쪽부터 지워야 인덱스가 안 꼬임
        print("5단계: 마커 및 템플릿 행 삭제...", file=sys.stderr)
        adjusted_end_row = end_row + len(items) - 1 if items else end_row
        ws.delete_rows(adjusted_end_row, 1)       # {/items}
        ws.delete_rows(template_row, 1)           # 템플릿 샘플
        ws.delete_rows(start_row, 1)              # {#items}

        # 6) 총합계 치환: {TOTAL_SUM} 셀을 찾아 SUM 수식으로 대체
        print("6단계: 총합계 수식 생성...", file=sys.stderr)
        data_last_row = first_data_row + len(items) - 1
        
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if cell.value == "{TOTAL_SUM}":
                    # 합계 열을 찾기 위해 데이터 행을 스캔
                    total_col = None
                    for col in range(1, max_cols + 1):
                        test_cell = ws.cell(row=first_data_row, column=col)
                        if isinstance(test_cell.value, (int, float)) and test_cell.value > 0:
                            total_col = col
                            break
                    
                    if total_col:
                        sum_formula = f"=SUM({openpyxl.utils.get_column_letter(total_col)}{first_data_row}:{openpyxl.utils.get_column_letter(total_col)}{data_last_row})"
                        cell.value = sum_formula
                        print(f"  총합계 수식 설정: {sum_formula}", file=sys.stderr)
                    else:
                        # 수식을 찾지 못한 경우 직접 계산
                        total_amount = sum(item.get('qty', 0) * item.get('unit_price', 0) for item in items)
                        cell.value = total_amount
                        print(f"  총합계 직접 계산: {total_amount}", file=sys.stderr)
                    
                    cell.font = Font(bold=True)
                    cell.number_format = "#,##0"

        # 7) 저장
        print(f"7단계: 파일 저장 중... {output_path}", file=sys.stderr)
        wb.save(output_path)
        
        return {"success": True, "output_path": str(output_path)}
        
    except Exception as e:
        print(f"오류 발생: {str(e)}", file=sys.stderr)
        return {"success": False, "error": str(e)}

def main():
    """메인 함수 - 명령줄 인자 처리"""
    parser = argparse.ArgumentParser(description='청구서 템플릿 렌더링')
    parser.add_argument('--template', required=True, help='템플릿 파일 경로')
    parser.add_argument('--output', required=True, help='출력 파일 경로')
    parser.add_argument('--data', required=True, help='JSON 데이터 (파일 경로 또는 JSON 문자열)')
    
    args = parser.parse_args()
    
    # JSON 데이터 로드
    try:
        if Path(args.data).exists():
            with open(args.data, 'r', encoding='utf-8') as f:
                payload = json.load(f)
        else:
            payload = json.loads(args.data)
    except Exception as e:
        print(json.dumps({"success": False, "error": f"JSON 데이터 파싱 오류: {str(e)}"}, ensure_ascii=False))
        sys.exit(1)
    
    # 템플릿 렌더링
    result = render_invoice(args.template, args.output, payload)
    print(json.dumps(result, ensure_ascii=False))

if __name__ == "__main__":
    # 예시 테스트 데이터
    if len(sys.argv) == 1:
        test_payload = {
            "invoice_no": "INV-2024-001",
            "issued_at": "2024-09-01",
            "client": "김철수",
            "project": "단독주택 신축",
            "site_addr": "서울시 강남구 역삼동 123-45",
            "header": {
                "client": "김철수",
                "project": "단독주택 신축",
                "site_addr": "서울시 강남구 역삼동 123-45",
                "issued_at": "2024-09-01"
            },
            "items": [
                {"title":"기초공사","desc":"건물 기초 및 지반 작업","spec":"토목공사","qty":1,"unit":"식","unit_price":3000000,"note":"콘크리트 강도 확인 필요"},
                {"title":"골조공사","desc":"철골 및 철근콘크리트 골조 작업","spec":"구조공사","qty":1,"unit":"식","unit_price":4000000,"note":"철근 간격 적정성 검토 필요"},
                {"title":"부대비용","desc":"자재운반 및 기타 부대비용","spec":"기타","qty":1,"unit":"식","unit_price":1500000,"note":"운반 차량 접근성 확인"}
            ]
        }
        print("테스트 실행 예시:", file=sys.stderr)
        print(f"python {sys.argv[0]} --template 청구서_템플릿.xlsx --output 청구서_결과.xlsx --data '{json.dumps(test_payload, ensure_ascii=False)}'", file=sys.stderr)
    else:
        main()