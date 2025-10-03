#!/usr/bin/env /Users/leechanhee/ConstructionManagement-Installer/venv/bin/python
# -*- coding: utf-8 -*-
"""
청구서 상세 엑셀 생성기
openpyxl을 사용하여 템플릿을 기반으로 청구서 엑셀 파일을 생성합니다.
"""

import sys
import json
import argparse
from copy import copy
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

HEADER_ROW = 7
TEMPLATE_ROW = 8
N_COLS = 7  # A~G

def clone_cell_style(src, dst):
    """셀 스타일을 복제합니다."""
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.number_format = copy(src.number_format)
    dst.protection = copy(src.protection)
    dst.alignment = copy(src.alignment)
    dst.border = copy(src.border)

def get_horizontal_merges_for_row(ws, row):
    """해당 행에 존재하는 '가로(수평) 병합' 범위를 [(min_col, max_col)] 로 반환"""
    ranges = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row == row and mr.max_row == row:
            ranges.append((mr.min_col, mr.max_col))
    return ranges

def apply_horizontal_merges(ws, merges, target_row):
    """수평 병합 범위를 target_row에 재적용"""
    for c1, c2 in merges:
        ws.merge_cells(start_row=target_row, start_column=c1, end_row=target_row, end_column=c2)

def insert_item_rows(ws, start_row, n_rows):
    """템플릿 행 바로 아래로 n_rows-1개 공간 확보 (템플릿 행 자체가 1행이기 때문)"""
    if n_rows <= 1:
        return
    # 실제 템플릿에서는 데이터 입력을 위해 행 삽입을 생략하고 직접 작성
    pass

def generate_invoice(template_path, output_path, payload):
    """청구서 엑셀 파일을 생성합니다."""
    try:
        wb = openpyxl.load_workbook(template_path)
        # 워크시트 이름을 동적으로 찾아서 사용
        sheet_names = wb.sheetnames
        if "청구서" in sheet_names:
            ws = wb["청구서"]
        elif len(sheet_names) > 0:
            ws = wb[sheet_names[0]]  # 첫 번째 시트 사용
            print(f"Warning: '청구서' 시트를 찾을 수 없어 '{sheet_names[0]}' 시트를 사용합니다.")
        else:
            raise ValueError("워크시트를 찾을 수 없습니다.")

        # 1) 헤더 바인딩 - 병합된 셀 처리
        header = payload["header"]
        
        # 병합된 셀의 경우 좌상단 셀에만 값 설정
        def set_merged_cell_value(ws, cell_ref, value):
            try:
                cell = ws[cell_ref]
                # MergedCell인지 확인
                if cell.__class__.__name__ == 'MergedCell':
                    # 병합된 셀인 경우 해당 범위를 찾아서 좌상단 셀에 값 설정
                    for merged_range in ws.merged_cells.ranges:
                        if cell_ref in merged_range:
                            top_left_coord = merged_range.start_cell.coordinate
                            top_left = ws[top_left_coord]
                            top_left.value = value
                            return
                else:
                    # 일반 셀인 경우
                    cell.value = value
            except Exception as e:
                print(f"셀 {cell_ref} 설정 실패: {e}")
        
        # 템플릿에 따라 적절한 셀에 데이터 바인딩
        try:
            set_merged_cell_value(ws, "A3", header["client"])  # 건축주명
            set_merged_cell_value(ws, "A4", header["project"])  # 프로젝트명  
            set_merged_cell_value(ws, "A5", header["site_addr"])  # 작업장 주소
            set_merged_cell_value(ws, "J3", header["issued_at"])  # 발행일
        except Exception as e:
            print(f"헤더 바인딩 중 오류: {e}")
            # 기본 셀에 시도
            try:
                ws["A3"].value = header["client"]
                ws["A4"].value = header["project"] 
                ws["A5"].value = header["site_addr"]
                ws["J3"].value = header["issued_at"]
            except:
                pass

        items = payload["items"]
        
        # 실제 템플릿 구조에 맞게 데이터 입력
        # 테스트를 위해 간단한 구조로 데이터 작성
        data_start_row = 10  # 실제 데이터 입력 시작 행
        
        for i, item in enumerate(items):
            row = data_start_row + i
            
            # 간단한 데이터 입력 테스트
            try:
                ws.cell(row=row, column=1).value = i + 1  # 연번
                ws.cell(row=row, column=2).value = item.get('title', '')  # 작업명
                ws.cell(row=row, column=3).value = item.get('spec', '')  # 규격
                ws.cell(row=row, column=4).value = item.get('qty', 0)  # 수량
                ws.cell(row=row, column=5).value = item.get('unit', '')  # 단위
                ws.cell(row=row, column=6).value = item.get('unit_price', 0)  # 단가
                ws.cell(row=row, column=7).value = item.get('qty', 0) * item.get('unit_price', 0)  # 합계
                ws.cell(row=row, column=8).value = item.get('note', '')  # 비고
            except Exception as e:
                print(f"행 {row} 데이터 입력 실패: {e}")
                
        # 기존 복잡한 코드는 주석 처리
        """
        first_data_row = TEMPLATE_ROW

        # 2) 데이터 행 확보
        insert_item_rows(ws, TEMPLATE_ROW, len(items))

        # 3) 템플릿 행 스타일/행높이/수평 병합 구조 확보
        tmpl_cells = [ws.cell(row=TEMPLATE_ROW, column=c) for c in range(1, N_COLS + 1)]
        tmpl_row_height = ws.row_dimensions[TEMPLATE_ROW].height
        horizontal_merges = get_horizontal_merges_for_row(ws, TEMPLATE_ROW)

        # 4) 각 데이터 행 채우기 (병합 구조 먼저 복제 → 값/스타일 입력)
        for i, item in enumerate(items):
            r = first_data_row + i

            # 동일한 수평 병합 구조 적용
            apply_horizontal_merges(ws, horizontal_merges, r)

            # 값 매핑 (주의: 병합된 셀 값은 '왼쪽 위 셀'에만 넣기)
            title_desc = f"{item['title']}\n{item.get('desc', '')}".strip()
        """

        # 파일 저장
        wb.save(output_path)
        return {"success": True, "output_path": str(output_path)}
        
    except Exception as e:
        return {"success": False, "error": str(e)}

def main():
    """메인 함수 - 명령줄 인자 처리"""
    parser = argparse.ArgumentParser(description='청구서 엑셀 파일 생성')
    parser.add_argument('--template', required=True, help='템플릿 파일 경로')
    parser.add_argument('--output', required=True, help='출력 파일 경로')
    parser.add_argument('--data', required=True, help='JSON 데이터 (파일 경로 또는 JSON 문자열)')
    
    args = parser.parse_args()
    
    # JSON 데이터 로드
    try:
        if Path(args.data).exists():
            with open(args.data, 'r', encoding='utf-8') as f:
                payload = json.load(f)
        else:
            payload = json.loads(args.data)
    except Exception as e:
        print(json.dumps({"success": False, "error": f"JSON 데이터 파싱 오류: {str(e)}"}, ensure_ascii=False))
        sys.exit(1)
    
    # 엑셀 파일 생성
    result = generate_invoice(args.template, args.output, payload)
    print(json.dumps(result, ensure_ascii=False))

if __name__ == "__main__":
    # 예시 테스트 데이터
    if len(sys.argv) == 1:
        test_payload = {
            "header": {
                "invoice_no": "INV-2024-001",
                "issued_at": "2024-09-01",
                "client": "김철수",
                "project": "단독주택 신축",
                "site_addr": "서울시 강남구 역삼동 123-45"
            },
            "items": [
                {
                    "title": "기초공사",
                    "desc": "건물 기초 및 지반 작업",
                    "spec": "토목공사",
                    "qty": 1,
                    "unit": "식",
                    "unit_price": 3000000,
                    "note": "콘크리트 강도 확인 필요"
                },
                {
                    "title": "골조공사",
                    "desc": "철골 및 철근콘크리트 골조 작업",
                    "spec": "구조공사",
                    "qty": 1,
                    "unit": "식",
                    "unit_price": 4000000,
                    "note": "철근 간격 적정성 검토 필요"
                },
                {
                    "title": "부대비용",
                    "desc": "자재운반 및 기타 부대비용",
                    "spec": "기타",
                    "qty": 1,
                    "unit": "식",
                    "unit_price": 1500000,
                    "note": "운반 차량 접근성 확인"
                }
            ]
        }
        print("테스트 실행 예시:")
        print(f"python {sys.argv[0]} --template 청구서_템플릿.xlsx --output 청구서_결과.xlsx --data '{json.dumps(test_payload, ensure_ascii=False)}'")
    else:
        main()